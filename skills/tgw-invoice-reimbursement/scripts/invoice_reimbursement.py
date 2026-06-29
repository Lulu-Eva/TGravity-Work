#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import shutil
import sys
from collections import Counter
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple


HEADERS = ["发票号码", "发票种类", "开票日期", "销售方", "货物、应税劳务及服务", "价税合计"]
DEFAULT_COMPANY = "深圳璐途文化科技有限公司"
DEFAULT_SHEET = "发票汇总"
PDF_EXT = ".pdf"
OFD_EXT = ".ofd"

RIDE_KEYWORDS = [
    "打车",
    "约车",
    "出租",
    "T3",
    "365",
    "曹操",
    "如祺",
    "风韵",
    "鞍马",
    "火箭",
    "小牛快跑",
    "迪尔",
    "高德打车",
    "及时用车",
    "喜行",
    "携华",
    "旅程",
    "添猫",
    "阳光",
    "东潮",
    "大象",
]

FLIGHT_KEYWORDS = ["机票", "携程", "华程", "代订机票", "代订附加", "航空"]

STATION_EN = {
    "Beijingbei": "北京北站",
    "Zhangjiakou": "张家口站",
    "Huhehaotedong": "呼和浩特东站",
    "Huhehaote": "呼和浩特站",
    "Tuoketuodong": "托克托东站",
    "Fuzhounan": "福州南站",
    "Shenzhenbei": "深圳北站",
    "Qingdao": "青岛站",
    "Beijingnan": "北京南站",
}


@dataclass
class ParsedInvoice:
    row: Dict[str, str]
    file: Path
    buyer: str = ""
    raw_text: str = ""
    confidence: str = "medium"
    issues: List[str] = field(default_factory=list)


@dataclass
class RunStats:
    pdf_count: int = 0
    ofd_count: int = 0
    skipped_ofd_count: int = 0
    parsed_count: int = 0
    written_count: int = 0
    skipped_duplicates: List[Tuple[str, str]] = field(default_factory=list)
    skipped_non_company: List[Tuple[str, str]] = field(default_factory=list)
    failed_files: List[Tuple[str, str]] = field(default_factory=list)
    railway_count: int = 0
    semantic_issues: List[Tuple[str, str, str]] = field(default_factory=list)
    empty_field_issues: List[Tuple[str, str]] = field(default_factory=list)
    amount_issues: List[Tuple[str, str]] = field(default_factory=list)
    low_confidence_company: List[Tuple[str, str]] = field(default_factory=list)
    cross_check_dirs: List[str] = field(default_factory=list)
    cross_table_duplicates: List[Tuple[str, str]] = field(default_factory=list)


def eprint(*parts: object) -> None:
    print(*parts, file=sys.stderr)


def compact(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def clean(value: str) -> str:
    text = compact(value)
    text = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", text)
    text = re.sub(r"(?<=[A-Za-z0-9])\s+(?=[A-Za-z0-9])", "", text)
    return re.sub(r"\s+", " ", text).strip(" ：:")


def squish(value: str) -> str:
    return re.sub(r"\s+", "", value or "")


def pick(pattern: str, text: str, default: str = "", flags: int = re.S) -> str:
    match = re.search(pattern, text, flags)
    return clean(match.group(1)) if match else default


def require_imports() -> Tuple[object, object, object, List[str]]:
    missing: List[str] = []
    try:
        from pdfminer.high_level import extract_text
    except Exception:
        extract_text = None
        missing.append("pdfminer.six")
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Border, Font, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        Workbook = load_workbook = Alignment = Border = Font = Side = get_column_letter = None
        missing.append("openpyxl")
    return extract_text, (Workbook, load_workbook), (Alignment, Border, Font, Side, get_column_letter), missing


def scan_files(invoice_dir: Path) -> Tuple[List[Path], int, int]:
    pdfs = sorted(path for path in invoice_dir.rglob("*") if path.is_file() and path.suffix.lower() == PDF_EXT)
    ofds = sorted(path for path in invoice_dir.rglob("*") if path.is_file() and path.suffix.lower() == OFD_EXT)
    pdf_stems = {path.stem for path in pdfs}
    skipped_ofd = sum(1 for path in ofds if path.stem in pdf_stems)
    return pdfs, len(ofds), skipped_ofd


def unique_output_path(base: Path) -> Path:
    if not base.exists():
        return base
    stem = base.stem
    suffix = base.suffix
    parent = base.parent
    index = 2
    while True:
        candidate = parent / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def normalize_date(value: str) -> str:
    text = clean(value)
    match = re.search(r"(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})", text)
    if not match:
        return text
    return f"{int(match.group(1)):04d}年{int(match.group(2)):02d}月{int(match.group(3)):02d}日"


def normalize_amount(value: str) -> str:
    text = clean(value).replace("¥", "").replace("￥", "").replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return ""
    try:
        return f"{Decimal(match.group(0)):.2f}"
    except InvalidOperation:
        return match.group(0)


def is_rail(text: str) -> bool:
    return "铁路电子客票" in text or "买票请到12306" in text or "中国铁路" in text


def has_flight_signal(filename: str, text: str, item: str = "") -> bool:
    haystack = f"{filename} {text} {item}"
    return any(keyword in haystack for keyword in FLIGHT_KEYWORDS)


def has_ride_signal(filename: str, text: str, item: str = "") -> bool:
    if has_flight_signal(filename, text, item):
        return False
    haystack = f"{filename} {text} {item}"
    return "旅客运输服务" in haystack or "客运服务费" in haystack or any(keyword in haystack for keyword in RIDE_KEYWORDS)


def extract_buyer(text: str, company_name: str) -> str:
    patterns = [
        r"购\s*买\s*方\s*信\s*息\s*名\s*称[：:\s]*(.*?)\s*统一社会(?:信用|信息)代码",
        r"购买方名称[：:\s]*(.*?)\s*统一社会信用代码",
        r"购买方[：:\s]*(.*?)\s*(?:纳税人识别号|统一社会信用代码|销\s*售\s*方)",
        r"名称[：:\s]*(.*?)\s*统一社会(?:信用|信息)代码",
    ]
    for pattern in patterns:
        value = pick(pattern, text)
        if value:
            return value
    return ""


def company_match_status(parsed: ParsedInvoice, company_name: str) -> str:
    if not company_name:
        return "matched"
    if parsed.buyer:
        return "matched" if company_name in parsed.buyer else "mismatch"
    if company_name in parsed.raw_text:
        return "low_confidence"
    return "missing"


def extract_seller(text: str) -> str:
    patterns = [
        r"销\s*售\s*方\s*信\s*息\s*名\s*称[：:\s]*(.*?)\s*统一社会(?:信用|信息)代码",
        r"销售方名称[：:\s]*(.*?)\s*统一社会信用代码",
        r"销售方[：:\s]*(.*?)\s*(?:纳税人识别号|统一社会信用代码|购买方)",
    ]
    for pattern in patterns:
        value = pick(pattern, text)
        if value:
            return value
    return ""


def extract_invoice_number(text: str, filename: str) -> str:
    patterns = [
        r"发\s*票\s*号\s*码[：:\s]*(\d{8,30})",
        r"发票号码[:：]\s*(\d{8,30})",
        r"\b(26\d{16,24})\b",
    ]
    for pattern in patterns:
        value = pick(pattern, text)
        if value:
            return value
    match = re.search(r"(\d{16,24})", filename)
    return match.group(1) if match else ""


def extract_invoice_date(text: str) -> str:
    patterns = [
        r"开\s*票\s*日\s*期[：:\s]*(\d{4}年\d{1,2}月\d{1,2}日)",
        r"开票日期[:：]\s*(\d{4}年\d{1,2}月\d{1,2}日)",
        r"(\d{4}年\d{1,2}月\d{1,2}日)",
    ]
    for pattern in patterns:
        value = pick(pattern, text)
        if value:
            return normalize_date(value)
    return ""


def extract_amount(text: str) -> str:
    patterns = [
        r"价税合计[（(]\s*大写\s*[）)].*?[（(]\s*小写\s*[）)]\s*[¥￥]?\s*([0-9]+(?:\.[0-9]+)?)",
        r"[（(]\s*小写\s*[）)]\s*[¥￥]?\s*([0-9]+(?:\.[0-9]+)?)",
        r"价税合计.*?[¥￥]\s*([0-9]+(?:\.[0-9]+)?)",
        r"合\s*计\s*[¥￥]?\s*([0-9]+(?:\.[0-9]+)?)",
    ]
    for pattern in patterns:
        value = pick(pattern, text)
        if value:
            return normalize_amount(value)
    return ""


def star_items(text: str) -> List[str]:
    normalized = clean(text.replace("\n", " "))
    candidates = re.findall(r"\*[^*￥¥\d]{2,24}\*[^￥¥\d\n]{2,40}", normalized)
    cleaned: List[str] = []
    for item in candidates:
        value = clean(item)
        value = re.split(r"(?:数量|单价|金额|税率|税额|合计|规格|型号)", value)[0]
        value = value.strip(" ;；,，")
        if value and value not in cleaned:
            cleaned.append(value)
    return cleaned


def extract_item_name(text: str, filename: str, seller: str) -> Tuple[str, List[str]]:
    issues: List[str] = []
    squished = squish(text)

    if "代订机票" in squished:
        return "*生产生活服务*代订机票产品", issues
    if "代订附加" in squished:
        return "*生产生活服务*代订附加产品", issues

    if "*交通运输服务*客运服务费" in squished:
        return "*交通运输服务*客运服务费", issues
    if "*运输服务*客运服务费" in squished:
        issues.append("客运服务费票面不是常见 *交通运输服务* 口径，请人工确认")
        return "*运输服务*客运服务费", issues

    items = star_items(text)
    if items:
        item = items[0]
        if "客运服务费" in item and "交通运输服务" not in item:
            issues.append("客运服务费项目名称不是 *交通运输服务*客运服务费")
        return item, issues

    if "顺丰" in seller or "收派服务" in text or "快递" in filename:
        return "*收派服务*收派服务费", ["顺丰/快递项目名称为候选修正，需人工确认"]
    if has_ride_signal(filename, text):
        return "*交通运输服务*客运服务费", ["客运服务费为规则补全，需核对票面原文"]
    return "", issues


def infer_invoice_type(filename: str, text: str, item: str) -> str:
    if is_rail(text):
        return "铁路电子客票"
    if has_flight_signal(filename, text, item):
        return "电子普通发票"
    if has_ride_signal(filename, text, item):
        return "出租车/网约车发票"
    if "增值税专用发票" in text:
        return "增值税专用发票"
    if "电子发票" in text or "电子普通发票" in text or "电 子 发 票" in text:
        return "电子普通发票"
    return "电子普通发票"


def extract_route(text: str) -> str:
    station_hits: List[str] = []
    for english, chinese in STATION_EN.items():
        if english in text and chinese not in station_hits:
            station_hits.append(chinese)
    chinese_hits = re.findall(r"([\u4e00-\u9fff]{2,12}站)", text)
    for station in chinese_hits:
        if station not in station_hits and not any(stop in station for stop in ["车站", "本站"]):
            station_hits.append(station)
    if len(station_hits) >= 2:
        return f"{station_hits[0]}-{station_hits[1]}"
    return ""


def parse_rail_ticket(path: Path, text: str, company_name: str) -> ParsedInvoice:
    invoice_number = extract_invoice_number(text, path.name)
    invoice_date = extract_invoice_date(text)
    buyer = extract_buyer(text, company_name)
    train = pick(r"\b([CGDZTKS]\d{1,5})\b", text)
    amount = normalize_amount(pick(r"票价[：:]\s*[￥¥]?\s*([0-9]+(?:\.[0-9]+)?)", text))
    service_type = "铁路运输服务"
    if "退票费" in text or "退票" in text:
        service_type = "退票费"
        refund_amount = normalize_amount(pick(r"退票费[：:]\s*[￥¥]?\s*([0-9]+(?:\.[0-9]+)?)", text))
        if refund_amount:
            amount = refund_amount
    elif "补票" in text:
        service_type = "补票"
    route = extract_route(text)
    item = " ".join(part for part in [service_type, route, train] if part)
    issues: List[str] = []
    if not route:
        issues.append("铁路票路线未能稳定识别")
    if not train and service_type != "退票费":
        issues.append("铁路票车次未能稳定识别")
    row = {
        "发票号码": invoice_number,
        "发票种类": "铁路电子客票",
        "开票日期": invoice_date,
        "销售方": "中国铁路",
        "货物、应税劳务及服务": item,
        "价税合计": amount,
    }
    return ParsedInvoice(row=row, file=path, buyer=buyer, raw_text=text, confidence="medium", issues=issues)


def parse_regular_invoice(path: Path, text: str, company_name: str) -> ParsedInvoice:
    buyer = extract_buyer(text, company_name)
    seller = extract_seller(text)
    item, issues = extract_item_name(text, path.name, seller)
    invoice_type = infer_invoice_type(path.name, text, item)
    row = {
        "发票号码": extract_invoice_number(text, path.name),
        "发票种类": invoice_type,
        "开票日期": extract_invoice_date(text),
        "销售方": seller,
        "货物、应税劳务及服务": item,
        "价税合计": extract_amount(text),
    }
    return ParsedInvoice(row=row, file=path, buyer=buyer, raw_text=text, confidence="medium", issues=issues)


def parse_pdf(path: Path, extract_text, company_name: str) -> Optional[ParsedInvoice]:
    try:
        text = extract_text(str(path)) or ""
    except Exception as exc:
        raise RuntimeError(f"PDF 文本抽取失败: {exc}") from exc
    if not clean(text):
        return ParsedInvoice(
            row={header: "" for header in HEADERS},
            file=path,
            raw_text="",
            confidence="low",
            issues=["PDF 无可抽取文本，第一版未启用 OCR"],
        )
    if is_rail(text):
        return parse_rail_ticket(path, text, company_name)
    return parse_regular_invoice(path, text, company_name)


def semantic_issues(parsed: ParsedInvoice) -> List[str]:
    row = parsed.row
    filename = parsed.file.name
    text = parsed.raw_text
    invoice_type = row.get("发票种类", "")
    seller = row.get("销售方", "")
    item = row.get("货物、应税劳务及服务", "")
    issues = list(parsed.issues)

    if has_flight_signal(filename, text, item) and invoice_type == "出租车/网约车发票":
        issues.append("机票/携程/华程疑似误判为网约车")
    if invoice_type == "出租车/网约车发票" and "客运服务费" in item and item != "*交通运输服务*客运服务费":
        issues.append("网约车项目名称不是常见 *交通运输服务*客运服务费")
    if invoice_type == "铁路电子客票":
        if seller != "中国铁路":
            issues.append("铁路票销售方不是中国铁路")
        if not (item.startswith("铁路运输服务") or item.startswith("补票") or item.startswith("退票费")):
            issues.append("铁路票项目名称异常")
    if "酒店" in seller and "住宿" not in item:
        issues.append("酒店发票项目名称未包含住宿")
    if "餐饮" in seller and "餐" not in item:
        issues.append("餐饮发票项目名称未包含餐")
    if any(bad in item for bad in ["无费", "None", "nan"]):
        issues.append("项目名称存在明显错误文本")
    return sorted(set(issue for issue in issues if issue))


def write_excel(rows: List[Dict[str, str]], output_path: Path, template: Optional[Path]):
    _, (Workbook, load_workbook), style_modules, missing = require_imports()
    if missing:
        raise SystemExit(f"missing dependencies: {', '.join(missing)}")
    Alignment, Border, Font, Side, get_column_letter = style_modules

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if template and template.exists():
        template_wb = load_workbook(template, data_only=False)
        template_ws = template_wb.active
        template_headers = [
            str(template_ws.cell(row=1, column=col).value or "").strip()
            for col in range(1, len(HEADERS) + 1)
        ]
        if template_headers != HEADERS:
            raise SystemExit(
                "template header mismatch: first row must be "
                + " | ".join(HEADERS)
            )
        shutil.copyfile(template, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        ws.title = DEFAULT_SHEET
        for row_index in range(ws.max_row, 1, -1):
            ws.delete_rows(row_index)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = DEFAULT_SHEET

    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    for row_index, row_data in enumerate(rows, 2):
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_index, column=col, value=row_data.get(header, ""))
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    for col, header in enumerate(HEADERS, 1):
        width = len(header) * 2 + 2
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            value = row_cells[0].value
            if value:
                width = max(width, sum(2 if ord(ch) > 127 else 1 for ch in str(value)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(width, 60)
    ws.freeze_panes = "A2"
    wb.save(output_path)


def read_invoice_numbers_from_xlsx(path: Path) -> List[str]:
    _, (_, load_workbook), _, missing = require_imports()
    if "openpyxl" in missing:
        return []
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return []
    ws = wb.active
    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
    try:
        invoice_col = headers.index("发票号码") + 1
    except ValueError:
        invoice_col = 1
    values = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=invoice_col).value
        if value:
            values.append(str(value).strip())
    return values


def collect_cross_check_dirs(output_dir: Path, template: Optional[Path], extra_dirs: Iterable[str]) -> List[Path]:
    dirs: List[Path] = []
    for candidate in [output_dir, template.parent if template else None]:
        if candidate and candidate.exists() and candidate not in dirs:
            dirs.append(candidate)
    for raw in extra_dirs:
        candidate = Path(raw).expanduser().resolve()
        if candidate.exists() and candidate.is_dir() and candidate not in dirs:
            dirs.append(candidate)
    return dirs


def cross_check(dirs: Iterable[Path], current_output: Path, invoice_numbers: Iterable[str]) -> List[Tuple[str, str]]:
    current = set(invoice_numbers)
    overlaps: List[Tuple[str, str]] = []
    seen_files = set()
    for directory in dirs:
        for path in sorted(directory.glob("发票填写*.xlsx")):
            resolved = path.resolve()
            if resolved == current_output.resolve() or resolved in seen_files:
                continue
            seen_files.add(resolved)
            prior_numbers = set(read_invoice_numbers_from_xlsx(path))
            for number in sorted(current & prior_numbers):
                overlaps.append((number, str(path)))
    return overlaps


def verify_rows(parsed_rows: List[ParsedInvoice], cross_check_dirs: List[Path], output_path: Path, stats: RunStats) -> None:
    counts = Counter(row.row.get("发票号码", "") for row in parsed_rows)
    for number, count in counts.items():
        if number and count > 1:
            stats.semantic_issues.append((number, "", f"表内重复 {count} 次"))
    for parsed in parsed_rows:
        number = parsed.row.get("发票号码", "")
        for header in HEADERS:
            if not str(parsed.row.get(header, "")).strip():
                stats.empty_field_issues.append((number or parsed.file.name, header))
        amount = parsed.row.get("价税合计", "")
        try:
            Decimal(str(amount))
        except InvalidOperation:
            stats.amount_issues.append((number or parsed.file.name, amount))
        for issue in semantic_issues(parsed):
            stats.semantic_issues.append((number or parsed.file.name, str(parsed.file), issue))
    stats.cross_check_dirs = [str(path) for path in cross_check_dirs]
    stats.cross_table_duplicates = cross_check(cross_check_dirs, output_path, [row.row.get("发票号码", "") for row in parsed_rows])


def write_report(report_path: Path, output_path: Path, rows: List[ParsedInvoice], stats: RunStats) -> None:
    total = Decimal("0")
    for parsed in rows:
        try:
            total += Decimal(parsed.row.get("价税合计", "0") or "0")
        except InvalidOperation:
            pass
    lines = [
        "# 发票报销复查报告",
        "",
        f"报销表：{output_path}",
        "",
        "## 汇总",
        "",
        f"- 原始 PDF：{stats.pdf_count}",
        f"- 原始 OFD：{stats.ofd_count}",
        f"- 跳过 OFD：{stats.skipped_ofd_count}",
        f"- 成功解析：{stats.parsed_count}",
        f"- 写入条数：{stats.written_count}",
        f"- 跳过重复：{len(stats.skipped_duplicates)}",
        f"- 跳过非公司抬头：{len(stats.skipped_non_company)}",
        f"- 铁路票：{stats.railway_count}",
        f"- 合计金额：{total:.2f}",
        "",
        "## 复查",
        "",
        f"- 空字段：{len(stats.empty_field_issues)}",
        f"- 金额异常：{len(stats.amount_issues)}",
        f"- 低置信公司抬头：{len(stats.low_confidence_company)}",
        f"- 项目名称/类型异常：{len(stats.semantic_issues)}",
        f"- 跨表重复：{len(stats.cross_table_duplicates)}",
        "",
        "## 跨表查重范围",
        "",
        *[f"- {path}" for path in stats.cross_check_dirs],
        "",
    ]
    sections = [
        ("跳过重复", stats.skipped_duplicates),
        ("跳过非公司抬头", stats.skipped_non_company),
        ("低置信公司抬头", stats.low_confidence_company),
        ("识别失败", stats.failed_files),
        ("空字段", stats.empty_field_issues),
        ("金额异常", stats.amount_issues),
        ("项目名称/类型异常", stats.semantic_issues),
        ("跨表重复", stats.cross_table_duplicates),
    ]
    for title, items in sections:
        if not items:
            continue
        lines.extend([f"## {title}", ""])
        for item in items:
            lines.append("- " + " | ".join(str(part) for part in item))
        lines.append("")
    report_path.write_text("\n".join(lines), encoding="utf-8")


def run(args: argparse.Namespace) -> int:
    extract_text, _, _, missing = require_imports()
    if missing:
        eprint("missing dependencies:", ", ".join(missing))
        eprint("install: python3 -m pip install --user openpyxl pdfminer.six")
        return 2

    invoice_dir = Path(args.invoice_dir).expanduser().resolve()
    if not invoice_dir.exists():
        eprint(f"invoice dir not found: {invoice_dir}")
        return 2
    date_label = args.date_label
    template = Path(args.template).expanduser().resolve() if args.template else None
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else (template.parent if template else Path.cwd())
    output_path = unique_output_path(output_dir / f"发票填写_{date_label}报销.xlsx")
    report_path = output_path.with_name(f"{output_path.stem}_复查报告.md")
    company_name = args.company_name or DEFAULT_COMPANY

    stats = RunStats()
    pdfs, stats.ofd_count, stats.skipped_ofd_count = scan_files(invoice_dir)
    stats.pdf_count = len(pdfs)
    if not pdfs:
        eprint(f"no PDF invoices found in {invoice_dir}")
        return 1

    parsed_rows: List[ParsedInvoice] = []
    seen = set()
    for pdf in pdfs:
        try:
            parsed = parse_pdf(pdf, extract_text, company_name)
        except Exception as exc:
            stats.failed_files.append((str(pdf), str(exc)))
            continue
        if not parsed:
            stats.failed_files.append((str(pdf), "解析结果为空"))
            continue
        number = parsed.row.get("发票号码", "").strip()
        if not number:
            stats.failed_files.append((str(pdf), "未识别到发票号码"))
            continue
        company_status = company_match_status(parsed, company_name)
        if company_status in {"mismatch", "missing"}:
            stats.skipped_non_company.append((number, str(pdf)))
            continue
        if company_status == "low_confidence":
            parsed.confidence = "low"
            parsed.issues.append("未稳定识别购买方字段，仅在全文中发现公司抬头，需人工确认")
            stats.low_confidence_company.append((number, str(pdf)))
        if number in seen:
            stats.skipped_duplicates.append((number, str(pdf)))
            continue
        seen.add(number)
        parsed_rows.append(parsed)
        stats.parsed_count += 1
        if parsed.row.get("发票种类") == "铁路电子客票":
            stats.railway_count += 1

    stats.written_count = len(parsed_rows)
    rows = [parsed.row for parsed in parsed_rows]
    write_excel(rows, output_path, template)
    cross_check_dirs = collect_cross_check_dirs(output_dir, template, args.cross_check_dir)
    verify_rows(parsed_rows, cross_check_dirs, output_path, stats)
    write_report(report_path, output_path, parsed_rows, stats)

    print(f"output: {output_path}")
    print(f"report: {report_path}")
    print(f"pdf_count: {stats.pdf_count}")
    print(f"written_count: {stats.written_count}")
    print(f"duplicates_skipped: {len(stats.skipped_duplicates)}")
    print(f"railway_count: {stats.railway_count}")
    print(f"issues: {len(stats.empty_field_issues) + len(stats.amount_issues) + len(stats.semantic_issues) + len(stats.cross_table_duplicates)}")
    return 0


def check(_: argparse.Namespace) -> int:
    _, _, _, missing = require_imports()
    if missing:
        print("status: missing")
        print("missing:", ", ".join(missing))
        print("install: python3 -m pip install --user openpyxl pdfminer.six")
        return 1
    print("status: ok")
    print("openpyxl: ok")
    print("pdfminer.six: ok")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="TGW invoice reimbursement organizer")
    sub = parser.add_subparsers(dest="command", required=True)
    check_parser = sub.add_parser("check", help="检查本机依赖")
    check_parser.set_defaults(func=check)

    run_parser = sub.add_parser("run", help="整理发票并生成报销 Excel")
    run_parser.add_argument("--invoice-dir", required=True, help="发票下载目录")
    run_parser.add_argument("--date-label", required=True, help="报销日期，例如 2026-06-21")
    run_parser.add_argument("--template", help="Excel 模板路径，可选")
    run_parser.add_argument("--output-dir", help="输出目录，可选；无模板时默认当前目录")
    run_parser.add_argument("--cross-check-dir", action="append", default=[], help="额外跨表查重目录，可重复传入")
    run_parser.add_argument("--company-name", default=DEFAULT_COMPANY, help="购买方公司抬头")
    run_parser.set_defaults(func=run)
    return parser


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
